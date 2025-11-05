#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generate a last-two-dates aligned TXT report from asx_eod_output/DailyData.csv.

Output columns:
  - Ticker (base code without ".AX")
  - Close (latest date, 3 dp)
  - +/- (change in close between last two dates, 3 dp)
  - % (percentage change versus previous close, 2 dp)
  - Units (shares on latest date, thousands separator)
  - Value (Units * Close_latest, 2 dp)
  - Day Gain (Units * (Close_latest - Close_prev), 2 dp)

Notes:
  - Only tickers listed in asx_eod_output/Tickers_and_Shares.txt are reported and
    are ordered as they appear in that file. If the holdings file is missing,
    tickers found in the CSV are used in alphabetical order.
  - The TOTAL row right-justifies only the Value and Day Gain columns, matching
    the alignment of the rows above.
"""

from __future__ import annotations

import csv
import os
from typing import Dict, List, Tuple, Optional

OUTPUT_DIR = "asx_eod_output"
CSV_PATH = os.path.join(OUTPUT_DIR, "DailyData.csv")
HOLDINGS_PATH = os.path.join(OUTPUT_DIR, "Tickers_and_Shares.txt")
REPORT_PATH = os.path.join(OUTPUT_DIR, "Report_LastTwoDates.txt")

# Dedicated subfolders for reports
REPORT_TXT_DIR = os.path.join(OUTPUT_DIR, "Report_TXT")
REPORT_CSV_DIR = os.path.join(OUTPUT_DIR, "Report_CSV")
REPORT_XLSX_DIR = os.path.join(OUTPUT_DIR, "Report_XLSX")


def _read_holdings(path: str) -> Optional[List[Tuple[str, int]]]:
    if not os.path.exists(path):
        return None
    out: List[Tuple[str, int]] = []
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#") or line.startswith("//"):
                continue
            if "," in line:
                parts = [p.strip() for p in line.split(",") if p.strip()]
            else:
                parts = line.split()
            if len(parts) < 2:
                continue
            ticker, shares_s = parts[0], parts[1]
            try:
                shares = int(float(shares_s))
            except ValueError:
                continue
            out.append((ticker, shares))
    return out


def _fmt_int(n: int) -> str:
    return f"{n:,}"


def _fmt_2(n: float) -> str:
    return f"{n:,.2f}"


def _fmt_3(n: float) -> str:
    # Do not force a leading '+'; negatives will show '-'.
    return f"{n:.3f}"


def _fmt_pct(n: float) -> str:
    # 2 dp with percent sign; negatives carry '-'.
    return f"{n:.2f}%"


def _base_code(ticker: str) -> str:
    return ticker.split(".")[0]


def read_csv_rows(path: str) -> List[Dict[str, str]]:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Cannot find CSV: {path}")
    rows: List[Dict[str, str]] = []
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)
    return rows


def parse_float(x: str) -> float:
    try:
        return float(x)
    except Exception:
        # Try to strip commas or quotes if any slipped in
        try:
            return float(x.replace(",", "").strip("\"'"))
        except Exception:
            return 0.0


def parse_int(x: str) -> int:
    try:
        return int(float(x))
    except Exception:
        try:
            return int(x.replace(",", "").strip("\"'"))
        except Exception:
            return 0


def last_two_dates(rows: List[Dict[str, str]]) -> List[str]:
    dates = sorted({r["Date"] for r in rows})
    if len(dates) < 2:
        return dates
    return dates[-2:]


def generate_last_two_report(csv_path: str = CSV_PATH,
                             holdings_path: str = HOLDINGS_PATH,
                             report_dir: Optional[str] = OUTPUT_DIR,
                             report_path: Optional[str] = None,
                             emit_csv: bool = True,
                             emit_excel: bool = True,
                             quiet: bool = False) -> Tuple[str, str, str]:
    """Generate the aligned last-two-dates report.

    Returns a tuple of the two dates used (oldest, newest) as strings.
    """
    rows = read_csv_rows(csv_path)
    if not rows:
        raise SystemExit("DailyData.csv is empty.")

    # Determine last two dates (ascending order returned)
    d1, d2 = last_two_dates(rows)
    if d1 == d2:
        raise SystemExit("Need at least two distinct dates in DailyData.csv")

    # Map (date, ticker) -> {close, shares}
    data: Dict[Tuple[str, str], Tuple[float, int]] = {}
    for r in rows:
        date_s = r["Date"]
        if date_s not in (d1, d2):
            continue
        t = r["Ticker"]
        close = parse_float(r.get("Close", "0"))
        shares = parse_int(r.get("Shares", "0"))
        data[(date_s, t)] = (close, shares)

    # Preferred order: holdings file order; fallback to tickers seen.
    holdings = _read_holdings(holdings_path)
    if holdings:
        tickers_in_order = [t for t, _ in holdings]
    else:
        tickers_in_order = sorted({r["Ticker"] for r in rows})

    # Build report rows
    report_rows: List[Dict[str, str]] = []
    total_value = 0.0
    total_gain = 0.0

    for ticker in tickers_in_order:
        if (d2, ticker) not in data:
            # Skip tickers that lack the latest date
            continue
        c2, units2 = data.get((d2, ticker), (0.0, 0))
        c1, _units1 = data.get((d1, ticker), (0.0, units2))
        change = c2 - c1
        pct = (change / c1 * 100.0) if c1 else 0.0
        value = units2 * c2
        day_gain = units2 * change

        total_value += value
        total_gain += day_gain

        report_rows.append({
            "Ticker": _base_code(ticker),
            "Close": _fmt_3(round(c2 + 1e-12, 3)),
            "+/-": _fmt_3(round(change + 1e-12, 3)),
            "%": _fmt_pct(round(pct + 1e-12, 2)),
            "Units": _fmt_int(units2),
            "Value": _fmt_2(round(value + 1e-12, 2)),
            "Day Gain": _fmt_2(round(day_gain + 1e-12, 2)),
        })

    # Compute column widths based on content + headers
    headers = ["Ticker", "Close", "+/-", "%", "Units", "Value", "Day Gain"]
    col_widths: Dict[str, int] = {}
    for h in headers:
        width = len(h)
        for r in report_rows:
            width = max(width, len(r[h]))
        col_widths[h] = width

    # Compose lines with right alignment for numeric columns
    gap = "  "  # two spaces between columns

    def fmt_row(row: Dict[str, str]) -> str:
        parts: List[str] = []
        for h in headers:
            text = row.get(h, "")
            width = col_widths[h]
            if h == "Ticker":
                parts.append(text.ljust(width))
            else:
                parts.append(text.rjust(width))
        return gap.join(parts)

    # Header
    lines: List[str] = []
    # Title row (do not alter column layout below)
    title = f"Portfolio value change between {d1} and {d2}"
    lines.append(title)
    # Blank line between title and table header
    lines.append("")
    title = {h: h for h in headers}
    lines.append(fmt_row(title))

    # Body
    for r in report_rows:
        lines.append(fmt_row(r))

    # TOTAL row: place "TOTAL" label, but anchor Value and Day Gain so that
    # their RIGHT edges match those of the data rows, even if that means
    # overlapping left columns.
    total_value_s = _fmt_2(round(total_value + 1e-12, 2))
    total_gain_s = _fmt_2(round(total_gain + 1e-12, 2))

    # Determine right-edge anchor positions for Value and Day Gain based on rows
    # (fallback to header if no rows, though in practice we will have rows).
    def right_edge_for(col: str) -> int:
        # Compute deterministic right edge from the column layout
        left = 0
        for h in headers:
            if h == col:
                break
            left += col_widths[h] + len(gap)
        return left + col_widths[col]

    value_right = right_edge_for("Value")
    gain_right = right_edge_for("Day Gain")

    # Establish a base line length large enough to hold everything
    base_len = max((len(x) for x in lines), default=0)
    base_len = max(base_len, value_right, gain_right)
    # Allow totals to extend further left if longer than any row value
    base_len = max(base_len, value_right, gain_right)

    buf = [" "] * base_len
    # Place TOTAL label starting at column 0 (left-aligned within ticker area)
    total_label = "TOTAL"
    for i, ch in enumerate(total_label):
        if i < len(buf):
            buf[i] = ch

    # Place Value total so its right edge matches the anchor
    v_start = max(0, value_right - len(total_value_s))
    for i, ch in enumerate(total_value_s):
        pos = v_start + i
        if pos >= len(buf):
            buf.extend([" "] * (pos + 1 - len(buf)))
        buf[pos] = ch

    # Place Day Gain total similarly
    g_start = max(0, gain_right - len(total_gain_s))
    for i, ch in enumerate(total_gain_s):
        pos = g_start + i
        if pos >= len(buf):
            buf.extend([" "] * (pos + 1 - len(buf)))
        buf[pos] = ch

    # Debug: ensure alignment (can be commented out later)
    # print(f"[DEBUG] value_right={value_right} gain_right={gain_right} v_start={v_start} g_start={g_start} base_len={len(buf)}")
    total_line = "".join(buf).rstrip()
    lines.append(total_line)

    # Prepare target directories
    base_dir = report_dir or os.path.dirname(csv_path) or OUTPUT_DIR
    txt_dir = REPORT_TXT_DIR if base_dir == OUTPUT_DIR else os.path.join(base_dir, "Report_TXT")
    csv_dir = REPORT_CSV_DIR if base_dir == OUTPUT_DIR else os.path.join(base_dir, "Report_CSV")
    xlsx_dir = REPORT_XLSX_DIR if base_dir == OUTPUT_DIR else os.path.join(base_dir, "Report_XLSX")

    # Ensure directories exist
    os.makedirs(txt_dir, exist_ok=True)
    if emit_csv:
        os.makedirs(csv_dir, exist_ok=True)
    if emit_excel:
        os.makedirs(xlsx_dir, exist_ok=True)

    # Choose the new filename format if not explicitly provided
    d2_compact = d2.replace("-", "")
    if report_path is None:
        report_path = os.path.join(txt_dir, f"Report_{d2_compact}.txt")
    else:
        # If a custom report_path is provided, honor it for TXT only.
        # Other formats still go to their dedicated subfolders under base_dir.
        txt_dir = os.path.dirname(report_path) or txt_dir

    with open(report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    # Emit CSV into dedicated folder
    csv_path_out = os.path.join(csv_dir, f"Report_{d2_compact}.csv")
    if emit_csv:
        # Numeric-friendly CSV (no thousands separators, percent as numeric string with %)
        headers = ["Ticker", "Close", "+/-", "%", "Units", "Value", "Day Gain"]
        # Build numeric-ish rows: keep percent as string with '%' to match visual report
        csv_rows: List[List[object]] = []
        for r in report_rows:
            # Strip formatting to numeric where applicable
            t = r["Ticker"]
            close = float(r["Close"]) if r["Close"] else ""
            chg = float(r["+/-"]) if r["+/-"] else ""
            pct = r["%"]  # keep as text with %
            units = int(r["Units"].replace(",", "")) if r["Units"] else 0
            val = float(r["Value"].replace(",", "")) if r["Value"] else 0.0
            gain = float(r["Day Gain"].replace(",", "")) if r["Day Gain"] else 0.0
            csv_rows.append([t, close, chg, pct, units, val, gain])
        # Append TOTAL row
        csv_rows.append(["TOTAL", "", "", "", "", round(total_value + 1e-12, 2), round(total_gain + 1e-12, 2)])
        with open(csv_path_out, "w", encoding="utf-8", newline="") as fcsv:
            writer = csv.writer(fcsv)
            writer.writerow([f"Portfolio value change between {d1} and {d2}"])
            writer.writerow([])
            writer.writerow(headers)
            writer.writerows(csv_rows)

    # Emit Excel into dedicated folder
    xlsx_path_out = os.path.join(xlsx_dir, f"Report_{d2_compact}.xlsx")
    excel_written = False
    if emit_excel:
        try:
            import pandas as pd  # type: ignore

            # Prepare DataFrame; keep percent as text
            df_rows = []
            for r in report_rows:
                df_rows.append({
                    "Ticker": r["Ticker"],
                    "Close": float(r["Close"]) if r["Close"] else None,
                    "+/-": float(r["+/-"]) if r["+/-"] else None,
                    "%": r["%"],
                    "Units": int(r["Units"].replace(",", "")) if r["Units"] else 0,
                    "Value": float(r["Value"].replace(",", "")) if r["Value"] else 0.0,
                    "Day Gain": float(r["Day Gain"].replace(",", "")) if r["Day Gain"] else 0.0,
                })
            # TOTAL row
            df_rows.append({
                "Ticker": "TOTAL", "Close": None, "+/-": None, "%": None,
                "Units": None, "Value": round(total_value + 1e-12, 2), "Day Gain": round(total_gain + 1e-12, 2)
            })
            df = pd.DataFrame(df_rows, columns=["Ticker", "Close", "+/-", "%", "Units", "Value", "Day Gain"])

            # Choose Excel engine: prefer openpyxl, else xlsxwriter
            engine = None
            try:
                import openpyxl  # type: ignore
                engine = "openpyxl"
            except Exception:
                try:
                    import xlsxwriter  # type: ignore
                    engine = "xlsxwriter"
                except Exception:
                    engine = None
            if engine is None:
                raise ImportError("Neither openpyxl nor xlsxwriter is installed")

            with pd.ExcelWriter(xlsx_path_out, engine=engine) as writer:  # type: ignore
                df.to_excel(writer, sheet_name="Report", index=False, startrow=2)

                if engine == "openpyxl":
                    from openpyxl.styles import Font
                    from openpyxl.utils import get_column_letter

                    ws = writer.sheets.get("Report")
                    if ws is None:
                        ws = writer.book["Report"]  # type: ignore[attr-defined]
                    # Title
                    ws.cell(row=1, column=1).value = f"Portfolio value change between {d1} and {d2}"
                    # Bold header
                    header_row = 3  # 1-based (startrow=2 -> header at row 3)
                    for cell in ws[header_row]:
                        cell.font = Font(bold=True)
                    # Number formats per column for all data rows (including TOTAL)
                    first_row = header_row + 1
                    last_row = header_row + len(df_rows)
                    for r_idx in range(first_row, last_row + 1):
                        ws.cell(row=r_idx, column=2).number_format = "0.000"     # Close
                        ws.cell(row=r_idx, column=3).number_format = "0.000"     # +/-
                        ws.cell(row=r_idx, column=5).number_format = "#,##0"     # Units
                        ws.cell(row=r_idx, column=6).number_format = "#,##0.00"  # Value
                        ws.cell(row=r_idx, column=7).number_format = "#,##0.00"  # Day Gain
                    # Bold TOTAL row
                    for cell in ws[last_row]:
                        cell.font = Font(bold=True)

                else:  # xlsxwriter
                    ws = writer.sheets["Report"]
                    wb = writer.book
                    # Title
                    ws.write(0, 0, f"Portfolio value change between {d1} and {d2}")
                    # Formats
                    header_fmt = wb.add_format({"bold": True})
                    num3_fmt = wb.add_format({"num_format": "0.000"})
                    int_fmt = wb.add_format({"num_format": "#,##0"})
                    money_fmt = wb.add_format({"num_format": "#,##0.00"})
                    bold_fmt = wb.add_format({"bold": True})
                    # Header row formatting (row index 2)
                    ws.set_row(2, None, header_fmt)
                    # Column formats
                    ws.set_column(1, 1, None, num3_fmt)   # Close
                    ws.set_column(2, 2, None, num3_fmt)   # +/-
                    ws.set_column(4, 4, None, int_fmt)    # Units
                    ws.set_column(5, 5, None, money_fmt)  # Value
                    ws.set_column(6, 6, None, money_fmt)  # Day Gain
                    # Bold TOTAL row
                    total_row_idx = 3 + len(df_rows) - 1  # zero-based row index
                    ws.set_row(total_row_idx, None, bold_fmt)

                excel_written = True
        except Exception as e:
            if not quiet:
                print(f"Warning: Excel file not written ({e})")

    if not quiet:
        print(f"Report written: {report_path}")
        print(f"Also wrote: {csv_path_out}")
        if emit_excel and excel_written:
            print(f"Also wrote: {xlsx_path_out}")
        print(f"Dates used: {d1} and {d2}")

    return d1, d2, report_path


def main() -> None:
    generate_last_two_report()


if __name__ == "__main__":
    main()
